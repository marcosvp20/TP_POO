import openpyxl
import string

#Classe para facilitar na hora de salvar o status dos itens da casa

class Planilha:
   
   #abre a planilha
   def __init__(self,nome_planilha) -> None:
      self.nome_planilha = nome_planilha
      self.workbook = openpyxl.load_workbook(self.nome_planilha)
      self.planilha = self.workbook['Sheet1']
      self.proxima_linha = self.planilha.max_row + 1
      
   #salva um lista de dados na planilha
   #PRÉ CONDIÇÃO: O nome precisa ser o primeiro na lista de dados
   def salvar(self, dados):
      if not self.verifica_se_objeto_existe(dados[0]):
         for i in range(0, len(dados)):
            self.planilha[f'{string.ascii_uppercase[i]}{self.proxima_linha}'] = dados[i]
         self.workbook.save(self.nome_planilha)
         self.exclui_linha_vazia()
   
   #edita o status do objeto
   #não edita o nome
   def editar(self, dados):
      i = 1
      for linha in self.planilha.iter_rows(min_row=1, values_only=True) :
         if linha[0] == dados[0]:
            for j in range(1, len(dados)):
               if linha[j] != dados[j]:
                  cell = self.planilha[f'{string.ascii_uppercase[j]}{i}']
                  cell.value = dados[j]
                  self.workbook.save(self.nome_planilha)
         i += 1

   #verifica se o objeto já está cadastrado
   def verifica_se_objeto_existe(self,nome):
      objetos_com_o_mesmo_nome = 0
      for linha in self.planilha.iter_rows(min_row=1, values_only=True) :
         if linha[0] == nome:
            objetos_com_o_mesmo_nome += 1
      if objetos_com_o_mesmo_nome > 0:
         return True
      return False
   
   #Retorna o valor da coluna desejada a partir do nome do objeto
   #O número das colunas começam em 1
   def retorna_valor(self, nome, coluna):
      for linha in self.planilha.iter_rows(min_row=1, values_only=True) :
         if linha[0] == nome:
            return linha[coluna-1]
   
   def exclui_linha_vazia(self):
      i = 1
      for linha in self.planilha.iter_rows(min_row=1, values_only=True):
         linhas_vazias = 0
         for j in range(0, len(linha)):
            if linha[j] == None:
               linhas_vazias += 1
         if linhas_vazias == len(linha):
            self.planilha.delete_rows(i)
            self.workbook.save(self.nome_planilha)
         i += 1
   
   def retorna_quantidade(self, tipo:str) ->  int:
      quantidade = 0
      for linha in self.planilha.iter_rows(min_row=1, values_only=True):
         if linha[1] == tipo:
            quantidade += 1
      return quantidade
