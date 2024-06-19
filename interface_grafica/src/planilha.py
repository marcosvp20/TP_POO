from interface_grafica.src.interfaces_src.IPlanilha import IPlanilha
import openpyxl
import string

class Planilha(IPlanilha):
   """
   Classe para facilitar na hora de salvar o status dos itens da casa.
   """

   def __init__(self, nome_planilha:str) -> None:
      """
      Inicializa a classe Planilha.

      Args:
         nome_planilha (str): O nome do arquivo da planilha.
      """
      self.nome_planilha = nome_planilha
      self.workbook = openpyxl.load_workbook(self.nome_planilha)
      self.planilha = self.workbook['Sheet1']
      self.proxima_linha = self.planilha.max_row + 1
      if not self.verifica_se_esta_vazio():
         self.exclui_linha_vazia()

   def salvar(self, dados:list) -> bool:
      """
      Salva uma lista de dados na planilha.

      Args:
         dados (list): A lista de dados a serem salvos.

      Returns:
         bool: True se os dados foram salvos com sucesso, False caso contrário.
      """
      if not self.verifica_se_objeto_existe(dados[0]):
         for i in range(0, len(dados)):
            self.planilha[f'{string.ascii_uppercase[i]}{self.proxima_linha}'] = dados[i]
         self.workbook.save(self.nome_planilha)
         self.exclui_linha_vazia()
         return True
      return False

   def editar(self, dados:list, coluna:int) -> None:
      """
      Edita o status do objeto na planilha.

      Args:
         dados (list): A lista de dados contendo o nome e o novo status do objeto.
         coluna (int): O número da coluna onde o status será atualizado.
      """
      i = 1
      for linha in self.planilha.iter_rows(min_row=1, values_only=True):
         if linha[0] == dados[0]:
            cell = self.planilha[f'{string.ascii_uppercase[coluna]}{i}']
            cell.value = dados[coluna]
            self.workbook.save(self.nome_planilha)
         i += 1

   def verifica_se_objeto_existe(self, nome: str) -> bool:
      """
      Verifica se o objeto já está cadastrado na planilha.

      Args:
         nome (str): O nome do objeto a ser verificado.

      Returns:
         bool: True se o objeto já existe na planilha, False caso contrário.
      """
      if not self.verifica_se_esta_vazio():
         for linha in self.planilha.iter_rows(min_row=1, values_only=True):
            if linha[0].upper() == nome.upper():
               return True
      return False

   def retorna_valor(self, nome:str, coluna:int) -> str:
      """
      Retorna o valor da coluna desejada a partir do nome do objeto.

      Args:
         nome (str): O nome do objeto.
         coluna (int): O número da coluna desejada.

      Returns:
         str: O valor da coluna do objeto.
      """
      for linha in self.planilha.iter_rows(min_row=1, values_only=True):
         if linha[0] == nome:
            return linha[coluna-1]

   def exclui_linha_vazia(self) -> None:
      """
      Exclui as linhas vazias da planilha.
      """
      i = 1
      for linha in self.planilha.iter_rows(min_row=1, values_only=True):
         linhas_vazias = 0
         for j in range(0, len(linha)):
            if linha[j] == None:
               linhas_vazias += 1
         if linhas_vazias == len(linha):
            self.planilha.delete_rows(i)
            self.workbook.save(self.nome_planilha)
         i += 1

   def retorna_quantidade(self, tipo:str) -> int: 
      """
      Retorna a quantidade da classe de objetos presente na planilha.

      Args:
         tipo (str): O tipo de objeto.

      Returns:
         int: A quantidade de objetos do tipo especificado.
      """
      quantidade = 0
      if not self.verifica_se_esta_vazio():
         for linha in self.planilha.iter_rows(min_row=1, values_only=True):
            if linha[1] == tipo:
               quantidade += 1
         return quantidade
      else:
         return 0

   def verifica_se_esta_vazio(self) -> bool:
      """
      Verifica se a planilha está vazia.

      Returns:
         bool: True se a planilha está vazia, False caso contrário.
      """
      for row in self.planilha.iter_rows():
         for cell in row:
            if cell.value is not None:
               return False
      return True

   def retorna_quantidade_linhas(self) -> int: 
      """
      Retorna a quantidade de dispositivos presentes na planilha.

      Returns:
         int: A quantidade de dispositivos presentes na planilha.
      """
      if self.verifica_se_esta_vazio():
         return 0
      return self.planilha.max_row


   def excluir_dispositivo(self, nome:str) -> bool:
      """
      Exclui um dispositivo da planilha.

      Args:
         nome (str): O nome do dispositivo a ser excluído.

      Returns:
         bool: True se o dispositivo foi excluído com sucesso, False caso contrário.
      """
      linha_a_excluir = 1
      for linha in self.planilha.iter_rows(min_row=1, values_only=True):
         if linha[0] == nome:
            self.planilha.delete_rows(linha_a_excluir)
            self.workbook.save(self.nome_planilha)
            return True
         linha_a_excluir += 1
      return False

   def retorna_coluna(self, letra_coluna:int) -> list:

         dados_coluna = []
         for celula in self.planilha[letra_coluna]:
            dados_coluna.append(celula.value)
      
         return dados_coluna
         
   def retorna_linha(self, numero_linha:int) -> list:
        """
        Retorna os valores de uma linha específica da planilha.

        Args:
            numero_linha (int): O número da linha a ser retornada.

        Returns:
            list: Uma lista contendo os valores das células da linha.
        """
        linha = self.planilha[numero_linha]
        valores_linha = [celula.value for celula in linha]

        return valores_linha

   def limpar_planilha(self) -> None:
      """
      Limpa todos os dados da planilha.
      """
      for linha in self.planilha.iter_rows():
         for celula in linha:
            celula.value = None
      self.planilha.delete_cols(1, self.planilha.max_column)
      self.planilha.delete_rows(1, self.planilha.max_row)
      self.workbook.save(self.nome_planilha)
   
   def excluir_linha(self, nome:str, coluna:int) -> bool:
      """
      Exclui a linha em que o parâmetro passado esta.

      Args:
         nome (str): O nome do dispositivo a ser excluído.
         coluna (int): coluna na qual o nome está

      Returns:
         bool: True se o dispositivo foi excluído com sucesso, False caso contrário.
      """
      linha_a_excluir = 1
      for linha in self.planilha.iter_rows(min_row=1, values_only=True):
         if linha[coluna] == nome:
            self.planilha.delete_rows(linha_a_excluir)
            self.workbook.save(self.nome_planilha)
            return True
         linha_a_excluir += 1
      return False
   
   def selecionar(self, objeto_id:str) -> None:
        """
        Seleciona um objeto para ser adicionado à uma automação.
        """
        for row in self.planilha.iter_rows(min_row=1, max_row=self.planilha.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == objeto_id:
                    self.planilha.cell(row=cell.row, column=6).value = True
                    self.workbook.save(self.nome_planilha)
                    break
