from src.planilha import Planilha
import string
import openpyxl

class PlanilhaAuto(Planilha):
    def __init__(self, nome_planilha) -> None:
        """
        Inicializa uma nova instância da classe PlanilhaAuto.

        Args:
            nome_planilha (str): O nome da planilha.
        """
        super().__init__(nome_planilha)
    
    def salvar(self, dados) -> None:
        """
        Salva várias automações com o mesmo nome na planilha.

        Args:
            dados (list): Uma lista contendo os dados a serem salvos na planilha.
        """
        for i in range(0, len(dados)):
            self.planilha[f'{string.ascii_uppercase[i]}{self.proxima_linha}'] = dados[i]
        self.workbook.save(self.nome_planilha)
        self.exclui_linha_vazia()

    
    def copia_planilha(self, planilha_origem:str) -> None:
        """
        Copia o conteúdo de uma planilha de origem para a planilha atual.

        Args:
            planilha_origem (str): O caminho da planilha de origem.
        """
        workbook_src = openpyxl.load_workbook(planilha_origem)
        sheet_src = workbook_src.active
        if sheet_src.max_row == 1 and sheet_src.max_column == 1 and sheet_src.cell(1, 1).value is None:
            return
        else:
            for row in sheet_src.iter_rows(values_only=True):
                self.planilha.append(row)
            
            self.workbook.save(self.nome_planilha)
            
    def adicionar_nome_primeira_celula(self, nome:str)->None:
        """
        Adiciona nome da automação na primeira coluna.
        """
        # Adicione self.nome no início da primeira célula de cada linha existente
        self.planilha.insert_cols(1)

        # Adicionar self.nome na nova primeira coluna de cada linha existente
        for linha in range(1, self.planilha.max_row + 1):
            self.planilha.cell(row=linha, column=1, value=nome)

        self.workbook.save(self.nome_planilha)
    
    def adiciona_coluna_de_selecao(self) -> None:
        """
        Adiciona uma coluna de seleção para confirmar a adição da configuração à automação.
        """
        if self.nome_planilha == 'planilhas/automacoestemp.xlsx':
            for row in range(1, self.planilha.max_row + 1):
                self.planilha.cell(row=row, column=6, value=False)
        
        self.workbook.save(self.nome_planilha)

    def excluir_coluna_de_selecao(self) -> None:
        """
        Exclui a coluna de seleção.
        """
        self.planilha.delete_cols(6)
        self.workbook.save(self.nome_planilha)

    